#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for
# each county

import openpyxl, pprint
from openpyxl.styles import Font
print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
countyData = {}

# Fill in countyData with each county's population and tracts.
print('Reading data...')
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    state  = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop    = sheet['D' + str(row)].value

    # Make sure the key for this state exists.
    countyData.setdefault(state, {})
    # Make sure the key for this county in this state exists.
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})

    # Each row represents one census tract, so increment by one.
    countyData[state][county]['tracts'] += 1
    # Increase the county pop by the pop in this census tract.
    countyData[state][county]['pop'] += int(pop)

# Creates a new excel sheet with parsed data
print('Updating excel sheet...')
wb.create_sheet(title="Parsed Data")
newSheet = wb['Parsed Data']
newSheet['A1'].font = Font(bold=True)
newSheet['A1'] = 'State'
newSheet['B1'].font = Font(bold=True)
newSheet['B1'] = 'County'
newSheet['C1'].font = Font(bold=True)
newSheet['C1'] = 'POP 2010'
newSheet['D1'].font = Font(bold=True)
newSheet['D1'] = '# Tracts'
newSheet['F2'].font = Font(bold=True)
newSheet['F2'] = 'States'
newSheet['G2'].font = Font(bold=True)
newSheet['G2'] = 'Pop Sums'

# Population of cells with data
i = 2
j = 3
for state in countyData:
    newSheet.cell(row=j, column=6).value = state # Separate area for state sums
    newSheet.cell(row=j, column=7).value = 0      
    for county in countyData[state]:
        newSheet.cell(row=i, column=1).value = state
        newSheet.cell(row=i, column=2).value = county
        newSheet.cell(row=i, column=3).value = countyData[state][county]['pop']
        newSheet.cell(row=i, column=4).value = countyData[state][county]['tracts']
        newSheet.cell(row=j, column=7).value = newSheet.cell(row=j, column=7).value + countyData[state][county]['pop']
        i += 1
    j += 1

newSheet.freeze_panes = 'A2'

print("Creating charts...")

azCounties = openpyxl.chart.Reference(newSheet, min_col=2, min_row=98, max_row=112)
azPops = openpyxl.chart.Reference(newSheet, min_col=3, min_row=98, max_row=112)

azChart = openpyxl.chart.PieChart()
azChart.title = 'Population Distribution in Arizona'
azChart.add_data(azPops, titles_from_data = False)
azChart.set_categories(azCounties)
azChart.legend.layout = openpyxl.chart.layout.Layout(
    openpyxl.chart.layout.ManualLayout(
        w=0.3
    )
)

usPops = openpyxl.chart.Reference(newSheet, min_col=7, min_row=3, max_row=53)
usStates = openpyxl.chart.Reference(newSheet, min_col=6, min_row=3, max_row=53)

usChart = openpyxl.chart.PieChart()
usChart.title = 'Population Distribution by State'
usChart.add_data(usPops, titles_from_data = False)
usChart.set_categories(usStates)
usChart.legend.layout = openpyxl.chart.layout.Layout(
    openpyxl.chart.layout.ManualLayout(
        w=0.5
    )
)
usChart.layout = openpyxl.chart.layout.Layout(
    openpyxl.chart.layout.ManualLayout(
        w=0.5
    )
)


newSheet.add_chart(azChart, 'I2')
newSheet.add_chart(usChart, 'I18')

print("Saving spreadsheet as updatedcensuspopdata.xlsx...")
wb.save('updatedcensuspopdata.xlsx')
print("Program finished.")